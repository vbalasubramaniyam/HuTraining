
import openpyxl
class ExcelUtil:

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
       # Users/vbalasubramaniyam
        book = openpyxl.load_workbook("/Users/vbalasubramaniyam/Documents/GitHub/HuTraining/Main_Sel_Assignment/Utils/Assignment.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                   # print[Dict]
        return[Dict]